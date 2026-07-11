#!/usr/bin/env python3
"""
build_workbook.py — render Welder Lots to an XLSX workbook.

Developed by the QCDatabase.AI team. Released as open source as part of the
Welder Lots System skill.

Output
------
A single .xlsx file containing:

  * A **cover sheet** ("Open Lots") listing every lot that needs attention
    (status = open), most-urgent first, with pass/fail tallies and how long the
    oldest weld in the lot has been waiting.
  * **One tab per welder** holding that welder's full weld history in
    chronological order, every row tagged with its lot name and lot status.

Lot math is delegated to ``welder_lots.py`` (the engine described in
``welder_lots_system_spec.md``). This module only marshals input and formats
the workbook, so the numbers are identical to any other consumer of the engine.

Input
-----
A JSON file (or stdin) shaped like:

    {
      "project": "Unit 40 Piping",              # optional cover label
      "generated_by": "QCDatabase.AI",          # optional cover label
      "settings": {"lot_perc": 5, "quals_required": 2,
                   "bust_lim": 2, "nde_types": ["RT"]},
      "welders": [
        {
          "wid": "12", "name": "John Doe", "start_date": null,
          "welds": [
            {
              "weld_number": "W-1001", "date_welded": "2026-01-05",
              "reweld_code": "", "joint": "BW", "loc": "shop", "wps": "P1",
              "size": "6", "sch": "40",
              "drawing": "U40-L12-S1", "package": "PKG-A", "spec": "A106",
              "nde": [{"type": "RT", "pass": true, "report": "R-500",
                       "date": "2026-01-10"}]
            }
          ]
        }
      ]
    }

A flat top-level ``"welds"`` list is also accepted; welds are grouped onto
welders by a ``welder_wid`` / ``welder`` / ``wid`` key.

Usage
-----
  python build_workbook.py --input data.json --output welder_lots.xlsx
  cat data.json | python build_workbook.py -o welder_lots.xlsx
"""

from __future__ import annotations

import argparse
import json
import sys
from collections import defaultdict
from datetime import datetime, timezone

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from welder_lots import LotSettings, compute_lots, OPEN, CLOSED


# --------------------------------------------------------------------------- #
# Styling constants                                                           #
# --------------------------------------------------------------------------- #
HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(bold=True, color="FFFFFF")
TITLE_FONT = Font(bold=True, size=14, color="1F3864")
SUB_FONT = Font(italic=True, color="595959")
OPEN_FILL = PatternFill("solid", fgColor="FCE4E4")      # light red
CLOSED_FILL = PatternFill("solid", fgColor="E2EFDA")    # light green
QUAL_FILL = PatternFill("solid", fgColor="FFF2CC")      # light amber
FAIL_FONT = Font(color="C00000", bold=True)
OPEN_FONT = Font(color="C00000", bold=True)
CLOSED_FONT = Font(color="375623", bold=True)
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP_TOP = Alignment(vertical="top", wrap_text=True)


# --------------------------------------------------------------------------- #
# Helpers                                                                      #
# --------------------------------------------------------------------------- #
def _welder_name(w: dict) -> str:
    if w.get("name"):
        return str(w["name"])
    parts = [w.get("first_name", ""), w.get("last_name", "")]
    name = " ".join(p for p in parts if p).strip()
    return name or str(w.get("wid", "Unknown"))


def _sanitize_sheet_title(base: str, used: set) -> str:
    """Excel sheet titles: <=31 chars, unique, none of []:*?/\\ ."""
    for ch in "[]:*?/\\":
        base = base.replace(ch, "-")
    base = base.strip().strip("'") or "Welder"
    base = base[:31]
    title = base
    n = 2
    while title.lower() in used:
        suffix = f" ({n})"
        title = base[:31 - len(suffix)] + suffix
        n += 1
    used.add(title.lower())
    return title


def _fmt_date(value) -> str:
    from welder_lots import _parse_dt
    dt = _parse_dt(value)
    return dt.strftime("%Y-%m-%d") if dt else (str(value) if value else "")


def _days_since(dt) -> str:
    if dt is None:
        return ""
    now = datetime.now(timezone.utc).replace(tzinfo=None)
    return str(max(0, (now - dt).days))


def _nde_cell(weld: dict, settings: LotSettings):
    """Return (types, reports, dates, result_text) for a weld's NDE reports."""
    reports = weld.get("nde") or weld.get("nde_reports") or []
    types, ids, dates = [], [], []
    passed = failed = False
    want = set(settings.nde_types)
    for r in reports:
        rtype = str(r.get("type") or r.get("nde_type") or "").upper()
        if want and rtype and rtype not in want:
            continue
        types.append(rtype)
        ids.append(str(r.get("report") or r.get("report_id") or ""))
        dates.append(_fmt_date(r.get("date")))
        pf = r.get("pass")
        if pf is None:
            pf = r.get("pass_fail")
        if bool(pf):
            passed = True
        else:
            failed = True
    result = "FAIL" if failed else ("PASS" if passed else "PENDING")
    return ("; ".join(types), "; ".join(ids), "; ".join(dates), result)


def _autosize(ws, max_width=48):
    widths = defaultdict(int)
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            longest = max((len(s) for s in str(cell.value).split("\n")), default=0)
            widths[cell.column] = max(widths[cell.column], longest)
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = min(max_width, w + 3)


# --------------------------------------------------------------------------- #
# Input normalization                                                          #
# --------------------------------------------------------------------------- #
def _load_welders(data: dict):
    """Return a list of welder dicts each carrying a ``welds`` list."""
    welders = list(data.get("welders") or [])
    by_key = {}
    for w in welders:
        w.setdefault("welds", [])
        key = str(w.get("wid") or _welder_name(w))
        by_key[key] = w

    # Fold a flat top-level welds list onto welders by welder key.
    for weld in data.get("welds") or []:
        key = str(weld.get("welder_wid") or weld.get("welder") or weld.get("wid") or "")
        if key not in by_key:
            holder = {"wid": key, "name": key, "welds": []}
            by_key[key] = holder
            welders.append(holder)
        by_key[key]["welds"].append(weld)
    return welders


# --------------------------------------------------------------------------- #
# Sheet writers                                                                #
# --------------------------------------------------------------------------- #
WELDER_COLUMNS = [
    ("Lot", 6), ("Status", 9), ("Weld #", 12), ("Date welded", 12),
    ("Joint", 7), ("Size", 6), ("Sch", 6), ("Loc", 8), ("WPS", 10),
    ("Drawing", 18), ("Package", 14), ("Spec", 10),
    ("NDE type", 10), ("Report", 12), ("NDE date", 12), ("Result", 9),
]


def _write_welder_sheet(ws, welder, welder_lots, settings):
    name = _welder_name(welder)
    wid = welder.get("wid", "")
    open_n = welder_lots.open_count

    ws["A1"] = f"{name}"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = (f"WID: {wid}    |    Lots: {len(welder_lots.lots)}    |    "
                f"Open lots: {open_n}")
    ws["A2"].font = SUB_FONT

    if welder_lots.error or not welder_lots.lots:
        ws["A4"] = welder_lots.error or "No lots found"
        ws["A4"].font = Font(italic=True, color="808080")
        return

    header_row = 4
    for c, (label, _) in enumerate(WELDER_COLUMNS, start=1):
        cell = ws.cell(row=header_row, column=c, value=label)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = BORDER

    r = header_row + 1
    for lot in welder_lots.lots:
        row_fill = (QUAL_FILL if lot.number == 0
                    else OPEN_FILL if lot.is_open else CLOSED_FILL)
        status_font = OPEN_FONT if lot.is_open else CLOSED_FONT
        for wrow in lot.welds:
            weld = wrow.source
            ndet, ndeid, nded, result = _nde_cell(weld, settings)
            values = [
                lot.label,
                lot.status.upper(),
                weld.get("weld_number", ""),
                _fmt_date(weld.get("date_welded")),
                str(weld.get("joint", "")).upper(),
                weld.get("size", ""),
                weld.get("sch", ""),
                weld.get("loc", ""),
                weld.get("wps", ""),
                weld.get("drawing", ""),
                weld.get("package", ""),
                str(weld.get("spec", "")).upper(),
                ndet, ndeid, nded, result,
            ]
            for c, v in enumerate(values, start=1):
                cell = ws.cell(row=r, column=c, value=v)
                cell.fill = row_fill
                cell.border = BORDER
                cell.alignment = WRAP_TOP
            ws.cell(row=r, column=2).font = status_font          # Status
            if result == "FAIL":
                ws.cell(row=r, column=16).font = FAIL_FONT       # Result
            r += 1

    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    _autosize(ws)


COVER_COLUMNS = ["Welder", "WID", "Lot", "Welds", "Pass", "Fail",
                 "Oldest weld", "Days waiting"]


def _write_cover(ws, project, generated_by, settings, welder_results):
    ws.title = "Open Lots"
    ws["A1"] = "Welder Lots — Open Lots Needing Attention"
    ws["A1"].font = TITLE_FONT
    stamp = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    bits = []
    if project:
        bits.append(f"Project: {project}")
    bits.append(f"Generated: {stamp}")
    if generated_by:
        bits.append(f"By: {generated_by}")
    ws["A2"] = "    |    ".join(bits)
    ws["A2"].font = SUB_FONT
    ws["A3"] = (f"Settings — Lot %: {settings.lot_perc:g}  |  Lot size: {settings.lot_size}  "
                f"|  Qualifiers: {settings.quals_required}  |  Fail limit: {settings.bust_lim}  "
                f"|  NDE: {', '.join(settings.nde_types)}")
    ws["A3"].font = SUB_FONT

    # Gather open lots across all welders.
    open_rows = []
    total_open = 0
    for welder, wl in welder_results:
        for lot in wl.open_lots:
            total_open += 1
            oldest = lot.oldest_date
            open_rows.append({
                "welder": _welder_name(welder),
                "wid": welder.get("wid", ""),
                "lot": lot.label,
                "welds": lot.size,
                "pass": lot.points,
                "fail": lot.strikes,
                "oldest": oldest,
                "days": _days_since(oldest),
            })
    # Most urgent first: longest-waiting, then most failures.
    open_rows.sort(key=lambda x: (-(int(x["days"]) if x["days"] else 0), -x["fail"]))

    ws["A5"] = (f"{len(welder_results)} welder(s)  |  {total_open} open lot(s) "
                f"across {sum(1 for _, wl in welder_results if wl.open_count)} welder(s)")
    ws["A5"].font = Font(bold=True, color=("C00000" if total_open else "375623"))

    header_row = 7
    if not open_rows:
        ws.cell(row=header_row, column=1,
                value="✓ All lots are closed. No welders need attention.").font = CLOSED_FONT
        _autosize(ws)
        return

    for c, label in enumerate(COVER_COLUMNS, start=1):
        cell = ws.cell(row=header_row, column=c, value=label)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = BORDER
        cell.alignment = Alignment(vertical="center")

    r = header_row + 1
    for row in open_rows:
        values = [row["welder"], row["wid"], row["lot"], row["welds"],
                  row["pass"], row["fail"],
                  row["oldest"].strftime("%Y-%m-%d") if row["oldest"] else "",
                  row["days"]]
        for c, v in enumerate(values, start=1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.fill = OPEN_FILL
            cell.border = BORDER
        ws.cell(row=r, column=3).font = OPEN_FONT  # Lot
        if row["fail"]:
            ws.cell(row=r, column=6).font = FAIL_FONT
        r += 1

    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    _autosize(ws)


# --------------------------------------------------------------------------- #
# Orchestration                                                                #
# --------------------------------------------------------------------------- #
def build_workbook(data: dict) -> Workbook:
    settings = LotSettings.from_dict(data.get("settings"))
    welders = _load_welders(data)

    # Compute lots for every welder up front (single source of truth).
    welder_results = []
    for welder in welders:
        wl = compute_lots(welder.get("welds", []), settings, welder)
        welder_results.append((welder, wl))

    wb = Workbook()
    cover = wb.active
    _write_cover(cover, data.get("project"), data.get("generated_by"),
                 settings, welder_results)

    used_titles = {"open lots"}
    for welder, wl in welder_results:
        title = _sanitize_sheet_title(str(welder.get("wid") or _welder_name(welder)),
                                      used_titles)
        ws = wb.create_sheet(title=title)
        _write_welder_sheet(ws, welder, wl, settings)

    return wb


def main(argv=None):
    parser = argparse.ArgumentParser(
        prog="build_workbook.py",
        description="Render Welder Lots to an XLSX workbook (cover sheet of open "
                    "lots + one tab per welder). Developed by the QCDatabase.AI team.",
    )
    parser.add_argument("-i", "--input", default=None,
                        help="Path to input JSON (default: read stdin).")
    parser.add_argument("-o", "--output", required=True,
                        help="Path to write the .xlsx workbook.")
    args = parser.parse_args(argv)

    if args.input:
        with open(args.input, "r", encoding="utf-8") as fh:
            data = json.load(fh)
    else:
        if sys.stdin.isatty():
            parser.error("no --input given and nothing on stdin")
        data = json.load(sys.stdin)

    wb = build_workbook(data)
    wb.save(args.output)

    n_welders = len(data.get("welders") or []) or len(_load_welders(data))
    open_total = sum(
        len(compute_lots(w.get("welds", []), LotSettings.from_dict(data.get("settings")), w).open_lots)
        for w in _load_welders(data)
    )
    print(f"Wrote {args.output}: {n_welders} welder tab(s), {open_total} open lot(s).",
          file=sys.stderr)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
